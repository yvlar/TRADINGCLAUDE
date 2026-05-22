from __future__ import annotations

import csv
import io
import logging

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import Response

from app.models.annotation import Annotation, AnnotationCreate
from app.services.annotation_service import AnnotationService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/annotations", tags=["annotations"])

_MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CSV_HEADERS = ["Ticker", "Analysis ID", "Note", "Créé le", "Mis à jour le"]
_XLSX_COL_WIDTHS = [10, 38, 40, 22, 22]


@router.post("", status_code=201, response_model=Annotation)
async def upsert_annotation(body: AnnotationCreate, request: Request) -> Annotation:
    """Crée ou remplace l'annotation pour un analysis_id (idempotent)."""
    service: AnnotationService = request.app.state.annotation_service
    try:
        return await service.upsert(body.analysis_id, body.note)
    except Exception as exc:
        logger.exception("Erreur lors de l'upsert annotation %s", body.analysis_id)
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/export.csv", summary="Export CSV des annotations", response_class=Response)
async def export_annotations_csv(request: Request) -> Response:
    """Exporte toutes les annotations en CSV avec le ticker correspondant."""
    service: AnnotationService = request.app.state.annotation_service
    rows = await service.get_all_with_ticker()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_CSV_HEADERS)
    for row in rows:
        created = row.get("created_at")
        updated = row.get("updated_at")
        writer.writerow([
            row.get("ticker") or "",
            row.get("analysis_id") or "",
            row.get("note") or "",
            created.isoformat() if created else "",
            updated.isoformat() if updated else "",
        ])

    # utf-8-sig ajoute automatiquement le BOM pour compatibilité Excel Windows
    csv_bytes = buf.getvalue().encode("utf-8-sig")
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=annotations.csv"},
    )


@router.get("/export.xlsx", summary="Export Excel des annotations", response_class=Response)
async def export_annotations_xlsx(request: Request) -> Response:
    """Exporte toutes les annotations en Excel avec le ticker correspondant."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    service: AnnotationService = request.app.state.annotation_service
    rows = await service.get_all_with_ticker()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annotations"

    ws.append(_CSV_HEADERS)

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        created = row.get("created_at")
        updated = row.get("updated_at")
        ws.append([
            row.get("ticker") or "",
            row.get("analysis_id") or "",
            row.get("note") or "",
            created.isoformat() if created else "",
            updated.isoformat() if updated else "",
        ])

    for i, width in enumerate(_XLSX_COL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    return Response(
        content=output.getvalue(),
        media_type=_MIME_XLSX,
        headers={"Content-Disposition": "attachment; filename=annotations.xlsx"},
    )


@router.get("/{analysis_id}", response_model=Annotation)
async def get_annotation(analysis_id: str, request: Request) -> Annotation:
    """Retourne l'annotation d'une analyse (404 si absente)."""
    service: AnnotationService = request.app.state.annotation_service
    annotation = await service.get(analysis_id)
    if annotation is None:
        raise HTTPException(status_code=404, detail=f"Aucune annotation pour {analysis_id}")
    return annotation
