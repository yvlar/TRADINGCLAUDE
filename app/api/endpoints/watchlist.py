from __future__ import annotations

import io
import logging
import uuid
from datetime import datetime, timezone

import redis.asyncio as aioredis
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response
from pydantic import BaseModel

from app.api.endpoints.admin import _require_admin
from app.models.watchlist import WatchlistCreate, WatchlistEntry
from app.orchestrator.core import AnalyzeRequest
from app.services.api_key_service import ApiKeyRecord
from app.services.email_service import EmailService
from app.services.esg_history_service import EsgHistoryService
from app.services.report import ReportService, _composite_alerte, _composite_label
from app.services.watchlist_pdf_service import WatchlistPdfService
from app.services.watchlist_service import WatchlistService
from app.utils.esg_utils import esg_verdict
from app.workers.tasks import run_full_analysis

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/watchlist", tags=["watchlist"])


class EsgThresholdUpdate(BaseModel):
    esg_alert_threshold: float


class PriceThresholdUpdate(BaseModel):
    threshold: float


class WatchlistEsgEntry(BaseModel):
    ticker: str
    last_esg_score: float | None
    esg_alert_threshold: float
    last_analyzed_at: datetime | None


class WatchlistEsgResponse(BaseModel):
    entries: list[WatchlistEsgEntry]

_MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLSX_HEADERS = ["Ticker", "Nom", "Date ajout", "Composite Score", "Label", "Alerte", "Notes", "Score ESG", "Verdict ESG", "Annotation"]
_XLSX_COL_WIDTHS = [10, 20, 12, 16, 10, 8, 30, 12, 14, 40]


# Alias retro-compatible (Sprint 88) — implementation deplacee dans app/utils/esg_utils.py
_esg_verdict = esg_verdict


def _generate_watchlist_xlsx(rows: list[dict]) -> bytes:
    """Génère un fichier Excel .xlsx depuis les données watchlist enrichies."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Watchlist"

    ws.append(_XLSX_HEADERS)

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        score = row.get("composite_score_latest")
        threshold = float(row.get("composite_alert_threshold") or 15.0)
        label = row.get("composite_label_latest") or _composite_label(score)
        alerte = _composite_alerte(score, threshold)
        created_at = row.get("created_at")
        date_ajout = created_at.strftime("%Y-%m-%d") if created_at else ""

        esg_score = row.get("last_esg_score")
        ws.append([
            row.get("ticker", ""),
            "",  # nom — champ absent du modèle actuel
            date_ajout,
            round(score, 1) if score is not None else None,
            label,
            alerte,
            "",  # notes — champ absent du modèle actuel
            round(esg_score, 1) if esg_score is not None else None,
            _esg_verdict(esg_score),
            row.get("derniere_annotation", ""),
        ])

    for i, width in enumerate(_XLSX_COL_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

_JOB_TTL = 86400  # 24 heures


@router.post("", status_code=201, response_model=WatchlistEntry)
async def add_entry(body: WatchlistCreate, request: Request) -> WatchlistEntry:
    """Ajoute un ticker à la watchlist persistante."""
    service: WatchlistService = request.app.state.watchlist_service
    return await service.add_entry(body)


@router.get("", response_model=list[WatchlistEntry])
async def list_entries(request: Request) -> list[WatchlistEntry]:
    """Liste tous les tickers suivis dans la watchlist."""
    service: WatchlistService = request.app.state.watchlist_service
    return await service.list_entries()


@router.delete("/{entry_id}", status_code=204)
async def delete_entry(entry_id: str, request: Request) -> Response:
    """Retire un ticker de la watchlist."""
    service: WatchlistService = request.app.state.watchlist_service
    deleted = await service.delete_entry(entry_id)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"Entrée {entry_id} introuvable")
    return Response(status_code=204)


@router.get(
    "/export.pdf",
    summary="Export PDF de la watchlist",
    response_class=Response,
)
async def export_watchlist_pdf(request: Request) -> Response:
    """Génère et retourne le rapport PDF de la watchlist complète."""
    service: WatchlistPdfService = request.app.state.watchlist_pdf_service
    composite_history_service = request.app.state.composite_history_service
    pool = request.app.state.db_pool
    try:
        pdf_bytes = await service.generate_watchlist_pdf(
            pool=pool,
            composite_history_service=composite_history_service,
        )
    except ValueError:
        raise HTTPException(status_code=404, detail="Watchlist vide — aucune position à exporter")
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="watchlist-{date_str}.pdf"'},
    )


@router.get(
    "/export.xlsx",
    summary="Export Excel de la watchlist",
    response_class=Response,
)
async def export_watchlist_xlsx(request: Request) -> Response:
    """Génère et retourne un fichier Excel avec toutes les positions de la watchlist."""
    service: WatchlistService = request.app.state.watchlist_service
    rows = await service.get_all_with_composite()
    content = _generate_watchlist_xlsx(rows)
    return Response(
        content=content,
        media_type=_MIME_XLSX,
        headers={"Content-Disposition": "attachment; filename=watchlist.xlsx"},
    )


@router.get("/esg-scores", response_model=WatchlistEsgResponse, summary="Scores ESG de la watchlist")
async def get_esg_scores(request: Request) -> WatchlistEsgResponse:
    """Retourne les entrées watchlist avec leurs scores ESG, triées par score DESC nulls en dernier."""
    service: WatchlistService = request.app.state.watchlist_service
    entries = await service.list_entries()
    sorted_entries = sorted(
        entries,
        key=lambda e: (e.last_esg_score is None, -(e.last_esg_score or 0.0)),
    )
    return WatchlistEsgResponse(
        entries=[
            WatchlistEsgEntry(
                ticker=e.ticker,
                last_esg_score=e.last_esg_score,
                esg_alert_threshold=e.esg_alert_threshold,
                last_analyzed_at=e.last_analyzed_at,
            )
            for e in sorted_entries
        ]
    )


@router.patch("/{entry_id}/esg-threshold", response_model=WatchlistEntry)
async def update_esg_threshold(
    entry_id: str,
    body: EsgThresholdUpdate,
    request: Request,
) -> WatchlistEntry:
    """Met à jour le seuil d'alerte ESG pour une entrée watchlist."""
    service: WatchlistService = request.app.state.watchlist_service
    entry = await service.get_entry(entry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Entrée {entry_id} introuvable")
    await service.update_esg_threshold(entry_id, body.esg_alert_threshold)
    updated = await service.get_entry(entry_id)
    return updated


@router.patch("/{entry_id}/price-threshold", response_model=WatchlistEntry)
async def update_price_threshold(
    entry_id: str,
    body: PriceThresholdUpdate,
    request: Request,
) -> WatchlistEntry:
    """Met à jour le seuil d'alerte de prix (en %) pour une entrée watchlist."""
    if body.threshold <= 0 or body.threshold > 100:
        raise HTTPException(status_code=422, detail="threshold doit être > 0 et <= 100")
    service: WatchlistService = request.app.state.watchlist_service
    entry = await service.get_entry(entry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Entrée {entry_id} introuvable")
    await service.update_price_threshold(entry_id, body.threshold / 100)
    updated = await service.get_entry(entry_id)
    return updated


@router.get("/{entry_id}/price-status", summary="Prix courant et écart vs valeur intrinsèque")
async def price_status(entry_id: str, request: Request) -> dict:
    """Retourne le prix courant, la valeur intrinsèque et l'écart pour une entrée watchlist."""
    service: WatchlistService = request.app.state.watchlist_service
    entry = await service.get_entry(entry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Entrée {entry_id} introuvable")

    yahoo = request.app.state.yahoo_extractor
    ratios = await yahoo.extract(entry.ticker)
    current_price: float = ratios.price

    intrinsic = entry.last_intrinsic_value
    ecart_pct: float | None = None
    alerte = False
    if intrinsic is not None and intrinsic != 0.0:
        ecart_pct = (current_price - intrinsic) / intrinsic
        alerte = abs(ecart_pct) >= entry.price_alert_threshold_pct

    return {
        "ticker": entry.ticker,
        "current_price": current_price,
        "intrinsic_value": intrinsic,
        "ecart_pct": ecart_pct,
        "alerte": alerte,
    }


@router.post("/report/send", summary="Envoie immédiatement le rapport PDF de la watchlist par email")
async def send_watchlist_report(request: Request) -> dict:
    """Génère un rapport PDF récapitulatif de la watchlist et l'envoie par email."""
    service: WatchlistService = request.app.state.watchlist_service
    entries = await service.list_entries()

    if not entries:
        return {"sent": False, "reason": "watchlist vide"}

    report_service = ReportService()
    pdf_bytes = report_service.generate_watchlist_summary_pdf(entries)

    import os
    from datetime import datetime, timezone

    to = os.environ.get("REPORT_EMAIL_TO", "")
    if not to:
        return {"sent": False, "reason": "REPORT_EMAIL_TO non configuré"}

    email_service = EmailService()
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    sent = await email_service.send_report(
        subject=f"Rapport watchlist — {len(entries)} position(s) — {date_str}",
        body_text=f"Rapport de votre watchlist : {len(entries)} position(s).\n\nCopilote Financier IA",
        pdf_bytes=pdf_bytes,
        filename=f"watchlist-{date_str}.pdf",
        to=to,
    )
    logger.info("Envoi immédiat rapport watchlist — sent=%s, positions=%d", sent, len(entries))
    return {"sent": sent}


@router.post(
    "/check-esg-degradation",
    summary="Vérification manuelle des dégradations ESG (admin)",
)
async def check_esg_degradation(
    request: Request,
    _admin: ApiKeyRecord | None = Depends(_require_admin),
) -> dict:
    """Déclenche la vérification de dégradation ESG pour toutes les entrées watchlist."""
    watchlist_service: WatchlistService = request.app.state.watchlist_service
    esg_history_service: EsgHistoryService = request.app.state.esg_history_service
    webhook_service = request.app.state.webhook_service
    slack_service = request.app.state.slack_service

    entries = await watchlist_service.list_entries()
    nb_alertes = 0

    for entry in entries:
        if entry.last_esg_score is None:
            continue
        try:
            previous_score = await esg_history_service.get_latest_previous(entry.ticker)
            if WatchlistService.check_esg_degradation(entry, previous_score):
                await webhook_service.send_esg_alert(
                    ticker=entry.ticker,
                    esg_score=entry.last_esg_score,
                    threshold=entry.esg_alert_threshold,
                )
                await slack_service.send_esg_alert(
                    ticker=entry.ticker,
                    score=entry.last_esg_score,
                    threshold=entry.esg_alert_threshold,
                )
                nb_alertes += 1
                logger.warning(
                    "Alerte dégradation ESG (manuel) — %s : score %.1f, seuil %.1f",
                    entry.ticker,
                    entry.last_esg_score,
                    entry.esg_alert_threshold,
                )
        except Exception:
            logger.exception("Erreur vérification dégradation ESG pour %s", entry.ticker)

    return {"triggered": nb_alertes}


@router.post("/{entry_id}/analyze", summary="Déclenche une analyse manuelle immédiate")
async def analyze_entry(entry_id: str, request: Request) -> dict:
    """Soumet une analyse Celery immédiate pour l'entrée watchlist donnée."""
    service: WatchlistService = request.app.state.watchlist_service
    entry = await service.get_entry(entry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Entrée {entry_id} introuvable")

    job_id = str(uuid.uuid4())
    r: aioredis.Redis = request.app.state.redis_pool
    analyze_request = AnalyzeRequest(
        ticker=entry.ticker,
        workflow=entry.workflow,
        ratios=entry.ratios,
    )
    request_dict = analyze_request.model_dump(mode="json")
    await r.set(f"job:{job_id}:status", "pending", ex=_JOB_TTL)
    run_full_analysis.delay(job_id, request_dict)
    logger.info(
        "Analyse manuelle watchlist — entry=%s, job=%s, ticker=%s",
        entry_id, job_id, entry.ticker,
    )
    return {"job_id": job_id}
