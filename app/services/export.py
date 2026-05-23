"""
Service d'export CSV/Excel pour les resultats screener.

Convertit un ScreenResult en fichier CSV ou Excel telechargeable.
Utilise le module csv integre (CSV) et openpyxl (Excel).
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone

logger = logging.getLogger(__name__)

_CSV_COLUMNS = [
    "ticker",
    "composite_score",
    "composite_label",
    "defensive_score",
    "verdict",
    "workflow",
    "cost_usd",
    "depuis_cache",
    "erreur",
]

_EXCEL_HEADER_STYLE = {
    "font_bold": True,
    "fill_color": "1F4E79",
    "font_color": "FFFFFF",
}

_LABEL_COLORS = {
    "FORT": "90EE90",
    "MODERE": "FFD700",
    "FAIBLE": "FF6B6B",
}


def _row_from_entry(entry) -> list:
    """Construit une ligne CSV depuis un ScreenEntry."""
    return [
        entry.ticker,
        f"{entry.composite_score:.1f}" if entry.composite_score is not None else "",
        entry.composite_label or "",
        str(entry.defensive_score) if entry.defensive_score is not None else "",
        entry.verdict or "",
        entry.workflow_utilise,
        f"{entry.cost_usd:.6f}",
        "oui" if entry.depuis_cache else "non",
        entry.erreur or "",
    ]


def export_to_csv(screen_result) -> bytes:
    """Convertit un ScreenResult en contenu CSV UTF-8."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    # En-tete avec metadonnees
    writer.writerow([f"# Export screener — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
    writer.writerow([f"# Tickers analyses: {screen_result.tickers_analyses}"])
    writer.writerow([f"# Workflow: {screen_result.workflow}"])
    writer.writerow([f"# Cout total: ${screen_result.cout_total_usd:.4f}"])
    writer.writerow([])  # ligne vide

    writer.writerow(_CSV_COLUMNS)
    for entry in screen_result.resultats:
        writer.writerow(_row_from_entry(entry))

    return output.getvalue().encode("utf-8-sig")  # BOM pour Excel francais


def export_to_excel(screen_result) -> bytes:
    """Convertit un ScreenResult en fichier Excel .xlsx."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        logger.error("openpyxl non installe — export Excel indisponible")
        raise RuntimeError("openpyxl requis pour l'export Excel. Installer avec: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screener"

    # Metadonnees
    ws.append([f"Export screener — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
    ws.append([f"Tickers analyses: {screen_result.tickers_analyses} | Workflow: {screen_result.workflow} | Cout: ${screen_result.cout_total_usd:.4f}"])
    ws.append([])

    header_row = 4
    ws.append(_CSV_COLUMNS)

    # Style de l'en-tete
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Donnees
    label_fills = {
        k: PatternFill(start_color=v, end_color=v, fill_type="solid")
        for k, v in _LABEL_COLORS.items()
    }

    for entry in screen_result.resultats:
        ws.append(_row_from_entry(entry))
        row = ws.max_row
        label = entry.composite_label
        if label and label.rstrip("E") in label_fills:  # FORT/MODERE/FAIBLE
            fill = label_fills.get(label, label_fills.get("MODERE"))
            if fill:
                ws.cell(row=row, column=2).fill = fill  # composite_score
                ws.cell(row=row, column=3).fill = fill  # composite_label

    # Largeurs de colonnes
    col_widths = [10, 16, 18, 16, 12, 22, 12, 14, 40]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = width

    # Figer l'en-tete
    ws.freeze_panes = f"A{header_row + 1}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
